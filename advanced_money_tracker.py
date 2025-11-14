import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import os
import hashlib
import logging
import json
from drive_sync import GoogleDriveSync

# Optional imports with fallbacks
try:
    from tkcalendar import DateEntry
    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as pdf_colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    plt.rcParams['font.family'] = 'DejaVu Sans'
    CHART_AVAILABLE = True
except ImportError:
    CHART_AVAILABLE = False

class JsonFormatter(logging.Formatter):
    def format(self, record):
        log_record = {
            "timestamp": datetime.utcnow().isoformat(),
            "level": record.levelname,
            "message": record.getMessage(),
            "filename": record.filename,
            "lineno": record.lineno,
            "function": record.funcName,
            "module": record.module,
        }
        if record.exc_info:
            log_record["exc_info"] = self.formatException(record.exc_info)
        if record.stack_info:
            log_record["stack_info"] = self.formatStack(record.stack_info)
        return json.dumps(log_record, ensure_ascii=False)

class AdvancedMoneyTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("💰 Advanced Money Tracker - সম্পূর্ণ সংস্করণ")
        self.root.geometry("1400x800")

        self.setup_logging()

        # Color Themes
        self.themes = {
            'light': {
                'bg': '#f5f5f5',
                'card': '#ffffff',
                'primary': '#1976D2',
                'secondary': '#424242',
                'success': '#4CAF50',
                'danger': '#F44336',
                'warning': '#FF9800',
                'text': '#212121',
                'text_secondary': '#757575'
            }
        }
        self.current_theme = 'light'
        self.colors = self.themes[self.current_theme]

        # Database
        self.db_path = os.getenv('MONEY_TRACKER_DB_PATH', 'advanced_money_tracker.db')
        self.init_database()
        self.drive_sync = GoogleDriveSync(self.db_path, self.logger)

        # Security Check
        if self.check_password_exists():
            if not self.show_login():
                self.root.destroy()
                return

        # GUI
        self.create_modern_gui()
        self.load_dashboard()

    def setup_logging(self):
        self.logger = logging.getLogger('MoneyTrackerApp')
        self.logger.setLevel(logging.INFO)

        # File handler
        log_file = 'money_tracker.log'
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(JsonFormatter())
        self.logger.addHandler(file_handler)

        # Console handler (optional, for development)
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.info("Application logging initialized.")

    def init_database(self):
        """সম্পূর্ণ ডেটাবেস সেটআপ"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()

        # Accounts
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS accounts (
                account_id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_name TEXT NOT NULL UNIQUE,
                account_type TEXT NOT NULL CHECK(account_type IN ('আয়', 'খরচ')),
                category TEXT,
                icon TEXT DEFAULT '💰',
                color TEXT DEFAULT '#1976D2',
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Payment Methods
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS payment_methods (
                method_id INTEGER PRIMARY KEY AUTOINCREMENT,
                method_name TEXT NOT NULL UNIQUE,
                method_type TEXT CHECK(method_type IN ('নগদ', 'ব্যাংক', 'মোবাইল ব্যাংকিং', 'কার্ড')),
                balance REAL DEFAULT 0,
                icon TEXT DEFAULT '💵',
                is_active INTEGER DEFAULT 1
            )
        ''')

        # Transactions
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                transaction_id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_date TEXT NOT NULL,
                transaction_time TEXT,
                account_id INTEGER NOT NULL,
                amount REAL NOT NULL CHECK(amount > 0),
                description TEXT,
                voucher_no TEXT UNIQUE,
                payment_method_id INTEGER,
                receipt_path TEXT,
                is_recurring INTEGER DEFAULT 0,
                tags TEXT,
                is_deleted INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (account_id) REFERENCES accounts(account_id),
                FOREIGN KEY (payment_method_id) REFERENCES payment_methods(method_id)
            )
        ''')

        # Budget
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS budget (
                budget_id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id INTEGER NOT NULL,
                month TEXT NOT NULL,
                budgeted_amount REAL NOT NULL,
                alert_threshold REAL DEFAULT 90,
                UNIQUE(account_id, month),
                FOREIGN KEY (account_id) REFERENCES accounts(account_id)
            )
        ''')

        # Recurring Templates
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS recurring_templates (
                template_id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id INTEGER NOT NULL,
                payment_method_id INTEGER,
                amount REAL NOT NULL,
                description TEXT,
                frequency TEXT CHECK(frequency IN ('দৈনিক', 'সাপ্তাহিক', 'মাসিক', 'বার্ষিক')),
                day_of_month INTEGER,
                is_active INTEGER DEFAULT 1,
                next_date TEXT,
                last_created TEXT,
                FOREIGN KEY (account_id) REFERENCES accounts(account_id)
            )
        ''')

        # Financial Goals
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS financial_goals (
                goal_id INTEGER PRIMARY KEY AUTOINCREMENT,
                goal_name TEXT NOT NULL,
                target_amount REAL NOT NULL,
                current_amount REAL DEFAULT 0,
                deadline TEXT,
                category TEXT,
                icon TEXT DEFAULT '🎯',
                is_achieved INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Security
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS security (
                id INTEGER PRIMARY KEY,
                password_hash TEXT,
                security_question TEXT,
                security_answer_hash TEXT
            )
        ''')

        # Initial Data
        self.cursor.execute('SELECT COUNT(*) FROM accounts')
        if self.cursor.fetchone()[0] == 0:
            accounts_data = [
                ('ঠিকাদারী আয়', 'আয়', 'ব্যবসায়িক', '💼', '#4CAF50'),
                ('অন্যান্য আয়', 'আয়', 'অন্যান্য', '💰', '#8BC34A'),
                ('দৈনিক বাজার', 'খরচ', 'পারিবারিক', '🛒', '#FF5722'),
                ('মুক্তার বাড়ী আনুষঙ্গিক খরচ', 'খরচ', 'পারিবারিক', '🏠', '#795548'),
                ('বাজারে বাসার আনুষঙ্গিক খরচ', 'খরচ', 'পারিবারিক', '🏡', '#9E9E9E'),
                ('ইউটিলিটি মুক্তার বাড়ি', 'খরচ', 'পারিবারিক', '⚡', '#FFC107'),
                ('ইউটিলিটি বাজারে বাসা', 'খরচ', 'পারিবারিক', '💡', '#FFEB3B'),
                ('ঠিকাদারী ব্যবসা', 'খরচ', 'ব্যবসায়িক', '🏗️', '#2196F3'),
                ('ঠিকাদারী লাইসেন্স রিনিউয়াল', 'খরচ', 'ব্যবসায়িক', '📋', '#03A9F4'),
                ('অফিসে স্টেশনারি', 'খরচ', 'ব্যবসায়িক', '📝', '#00BCD4'),
                ('রাজনৈতিক খরচ', 'খরচ', 'সামাজিক', '🗳️', '#9C27B0'),
                ('অনুদান', 'খরচ', 'সামাজিক', '🤝', '#E91E63'),
                ('মেয়র স্যারের খরচ', 'খরচ', 'সামাজিক', '👔', '#F44336'),
                ('রছি ভাইয়ের খরচ', 'খরচ', 'সামাজিক', '👨', '#FF5722'),
                ('বেতন', 'খরচ', 'নিয়মিত', '💼', '#607D8B'),
                ('যাতায়াত খরচ', 'খরচ', 'নিয়মিত', '🚗', '#3F51B5'),
                ('নাস্তা/আপ্যায়ন', 'খরচ', 'নিয়মিত', '☕', '#FF9800')
            ]
            self.cursor.executemany(
                'INSERT INTO accounts (account_name, account_type, category, icon, color) VALUES (?, ?, ?, ?, ?)',
                accounts_data
            )
            # Default payment methods
            payment_methods = [
                ('নগদ টাকা', 'নগদ', 0, '💵'),
                ('ব্যাংক অ্যাকাউন্ট', 'ব্যাংক', 0, '🏦'),
                ('bKash', 'মোবাইল ব্যাংকিং', 0, '📱'),
                ('Nagad', 'মোবাইল ব্যাংকিং', 0, '💳'),
                ('Rocket', 'মোবাইল ব্যাংকিং', 0, '🚀')
            ]
            self.cursor.executemany(
                'INSERT INTO payment_methods (method_name, method_type, balance, icon) VALUES (?, ?, ?, ?)',
                payment_methods
            )
            self.conn.commit()

    def is_numeric(self, value):
        try:
            float(value)
            return True
        except ValueError:
            return False

    def check_password_exists(self):
        """পাসওয়ার্ড চেক"""
        self.cursor.execute('SELECT password_hash FROM security WHERE id = 1')
        result = self.cursor.fetchone()
        return result is not None and result[0] is not None

    def show_login(self):
        """লগইন স্ক্রিন"""
        login_window = tk.Toplevel(self.root)
        login_window.title("🔐 লগইন")
        login_window.geometry("400x250")
        login_window.transient(self.root)
        login_window.grab_set()

        x = (login_window.winfo_screenwidth() // 2) - 200
        y = (login_window.winfo_screenheight() // 2) - 125
        login_window.geometry(f"400x250+{x}+{y}")

        tk.Label(
            login_window,
            text="🔐 পাসওয়ার্ড প্রবেশ করুন",
            font=('Segoe UI', 16, 'bold')
        ).pack(pady=30)

        password_entry = tk.Entry(login_window, font=('Segoe UI', 14), show='●', width=20)
        password_entry.pack(pady=20)
        password_entry.focus()

        result = [False]

        def verify_password():
            password = password_entry.get()
            password_hash = hashlib.sha256(password.encode()).hexdigest()
            self.cursor.execute('SELECT password_hash FROM security WHERE id = 1')
            stored_hash = self.cursor.fetchone()[0]

            if password_hash == stored_hash:
                result[0] = True
                login_window.destroy()
            else:
                messagebox.showerror("ত্রুটি", "ভুল পাসওয়ার্ড!")
                password_entry.delete(0, tk.END)

        tk.Button(
            login_window,
            text="✅ লগইন",
            font=('Segoe UI', 12, 'bold'),
            bg='#4CAF50',
            fg='white',
            padx=30,
            pady=10,
            bd=0,
            cursor='hand2',
            command=verify_password
        ).pack(pady=10)

        password_entry.bind('<Return>', lambda e: verify_password())
        login_window.protocol("WM_DELETE_WINDOW", lambda: login_window.destroy())
        login_window.wait_window()
        return result[0]

    def create_modern_gui(self):
        """Modern GUI"""
        self.root.configure(bg=self.colors['bg'])

        # Top Bar
        self.nav_bar = tk.Frame(self.root, bg=self.colors['primary'], height=60)
        self.nav_bar.pack(fill=tk.X, side=tk.TOP)
        self.nav_bar.pack_propagate(False)

        tk.Label(
            self.nav_bar,
            text="💰 Advanced Money Tracker",
            font=('Segoe UI', 18, 'bold'),
            bg=self.colors['primary'],
            fg='white'
        ).pack(side=tk.LEFT, padx=20)

        # Main Container
        self.main_container = tk.Frame(self.root, bg=self.colors['bg'])
        self.main_container.pack(fill=tk.BOTH, expand=True)

        # Sidebar
        self.sidebar = tk.Frame(self.main_container, bg=self.colors['card'], width=220)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        # Content Area
        self.content_area = tk.Frame(self.main_container, bg=self.colors['bg'])
        self.content_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Menu Items
        menu_items = [
            ("🏠 ড্যাশবোর্ড", self.show_dashboard),
            ("➕ নতুন এন্ট্রি", self.show_new_entry),
            ("📊 রিপোর্ট", self.show_reports),
            ("💰 বাজেট", self.show_budget),
            ("🔄 Recurring", self.show_recurring),
            ("🎯 লক্ষ্য", self.show_goals),
            ("💳 অ্যাকাউন্ট", self.show_payment_methods),
            ("📤 এক্সপোর্ট", self.show_export),
            ("🔍 খুঁজুন", self.show_search),
            ("⚙️ সেটিংস", self.show_settings)
        ]

        self.menu_label = tk.Label(
            self.sidebar,
            text="মেনু",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        self.menu_label.pack(fill=tk.X, padx=15, pady=15)

        self.menu_buttons = []
        for text, command in menu_items:
            btn = tk.Button(
                self.sidebar,
                text=text,
                font=('Segoe UI', 11),
                bg=self.colors['card'],
                fg=self.colors['text'],
                bd=0,
                anchor='w',
                padx=20,
                pady=12,
                cursor='hand2',
                command=command
            )
            btn.pack(fill=tk.X, padx=5, pady=2)
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg=self.colors['primary'], fg='white'))
            btn.bind('<Leave>', lambda e, b=btn: b.config(bg=self.colors['card'], fg=self.colors['text']))
            self.menu_buttons.append(btn)

    def clear_content_area(self):
        for widget in self.content_area.winfo_children():
            widget.destroy()

    def show_dashboard(self):
        """Dashboard - সম্পূর্ণ"""
        self.clear_content_area()

        # Header
        header = tk.Frame(self.content_area, bg=self.colors['bg'])
        header.pack(fill=tk.X, pady=(0, 20))

        tk.Label(
            header,
            text="📊 ড্যাশবোর্ড",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(side=tk.LEFT)

        tk.Label(
            header,
            text=datetime.now().strftime("%d %B, %Y"),
            font=('Segoe UI', 12),
            bg=self.colors['bg'],
            fg=self.colors['text_secondary']
        ).pack(side=tk.RIGHT)

        # Summary Cards
        cards_frame = tk.Frame(self.content_area, bg=self.colors['bg'])
        cards_frame.pack(fill=tk.X, pady=10)

        today = datetime.now().strftime('%Y-%m-%d')
        month_start = datetime.now().replace(day=1).strftime('%Y-%m-%d')

        # Get data
        self.cursor.execute('''
            SELECT COALESCE(SUM(t.amount), 0)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            WHERE t.transaction_date = ? AND a.account_type = 'আয়' AND t.is_deleted = 0
        ''')
        today_income = self.cursor.fetchone()[0]

        self.cursor.execute('''
            SELECT COALESCE(SUM(t.amount), 0)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            WHERE t.transaction_date = ? AND a.account_type = 'খরচ' AND t.is_deleted = 0
        ''')
        today_expense = self.cursor.fetchone()[0]

        self.cursor.execute('''
            SELECT COALESCE(SUM(t.amount), 0)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            WHERE t.transaction_date BETWEEN ? AND ? AND a.account_type = 'আয়' AND t.is_deleted = 0
        ''')
        month_income = self.cursor.fetchone()[0]

        self.cursor.execute('''
            SELECT COALESCE(SUM(t.amount), 0)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            WHERE t.transaction_date BETWEEN ? AND ? AND a.account_type = 'খরচ' AND t.is_deleted = 0
        ''')
        month_expense = self.cursor.fetchone()[0]

        self.cursor.execute('SELECT COALESCE(SUM(balance), 0) FROM payment_methods WHERE is_active = 1')
        total_balance = self.cursor.fetchone()[0]

        cards_data = [
            ("আজকের আয়", f"৳ {today_income:,.2f}", "#4CAF50", "💰"),
            ("আজকের খরচ", f"৳ {today_expense:,.2f}", "#F44336", "💸"),
            ("মাসিক আয়", f"৳ {month_income:,.2f}", "#2196F3", "📈"),
            ("মাসিক খরচ", f"৳ {month_expense:,.2f}", "#FF9800", "📉"),
            ("মোট ব্যালেন্স", f"৳ {total_balance:,.2f}", "#9C27B0", "💵")
        ]

        for idx, (title, value, color, icon) in enumerate(cards_data):
            card = tk.Frame(cards_frame, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
            card.grid(row=idx//3, column=idx%3, padx=10, pady=10, sticky='ew')
            cards_frame.grid_columnconfigure(idx%3, weight=1)

            tk.Label(card, text=icon, font=('Segoe UI', 24), bg=self.colors['card']).pack(pady=(15, 5))
            tk.Label(card, text=title, font=('Segoe UI', 11), bg=self.colors['card'], fg=self.colors['text_secondary']).pack()
            tk.Label(card, text=value, font=('Segoe UI', 18, 'bold'), bg=self.colors['card'], fg=color).pack(pady=(5, 15))

        # Recent Transactions
        recent_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        recent_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))

        tk.Label(
            recent_frame,
            text="📋 সাম্প্রতিক লেনদেন",
            font=('Segoe UI', 14, 'bold'),
            bg=self.colors['card'],
            fg=self.colors['text']
        ).pack(anchor='w', padx=15, pady=15)

        tree_frame = tk.Frame(recent_frame, bg=self.colors['card'])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))

        columns = ('তারিখ', 'খাত', 'পরিমাণ', 'পেমেন্ট')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=8)

        for col in columns:
            tree.heading(col, text=col)
            tree.column('তারিখ', width=100)
            tree.column('খাত', width=300)
            tree.column('পরিমাণ', width=150)
            tree.column('পেমেন্ট', width=150)

        self.cursor.execute('''
            SELECT t.transaction_date, a.account_name, t.amount, COALESCE(pm.method_name, 'N/A')
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            LEFT JOIN payment_methods pm ON t.payment_method_id = pm.method_id
            WHERE t.is_deleted = 0
            ORDER BY t.transaction_date DESC, t.transaction_time DESC
            LIMIT 10
        ''')
        for row in self.cursor.fetchall():
            tree.insert('', 'end', values=(row[0], row[1], f"৳ {row[2]:,.2f}", row[3]))

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def show_new_entry(self):
        """নতুন এন্ট্রি - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="➕ নতুন লেনদেন",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        form_card = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        form_card.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Form
        form_container = tk.Frame(form_card, bg=self.colors['card'])
        form_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)

        # Type
        tk.Label(form_container, text="ধরন:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=0, column=0, sticky='w', pady=10)
        trans_type = tk.StringVar(value='খরচ')
        type_frame = tk.Frame(form_container, bg=self.colors['card'])
        type_frame.grid(row=0, column=1, sticky='w', pady=10)
        tk.Radiobutton(type_frame, text="💰 আয়", variable=trans_type, value='আয়', bg=self.colors['card']).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(type_frame, text="💸 খরচ", variable=trans_type, value='খরচ', bg=self.colors['card']).pack(side=tk.LEFT, padx=10)

        # Date
        tk.Label(form_container, text="তারিখ:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=1, column=0, sticky='w', pady=10)
        date_entry = tk.Entry(form_container, font=('Segoe UI', 11), width=30)
        date_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        date_entry.grid(row=1, column=1, sticky='w', pady=10)

        # Account
        tk.Label(form_container, text="খাত:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=2, column=0, sticky='w', pady=10)
        account_var = tk.StringVar()
        account_combo = ttk.Combobox(form_container, textvariable=account_var, values=[], width=38, state='readonly')
        account_combo.grid(row=2, column=1, sticky='w', pady=10)
        
        # Function to update accounts based on transaction type
        def update_accounts(*args):
            selected_type = trans_type.get()
            self.cursor.execute('SELECT account_name FROM accounts WHERE account_type = ? AND is_active = 1', (selected_type,))
            updated_accounts = [r[0] for r in self.cursor.fetchall()]
            account_combo['values'] = updated_accounts
            if updated_accounts:
                account_combo.current(0)
            else:
                account_var.set("")

        trans_type.trace_add('write', update_accounts)
        update_accounts() # Initial call to populate accounts based on default trans_type

        # Amount
        tk.Label(form_container, text="পরিমাণ (৳):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=3, column=0, sticky='w', pady=10)
        amount_entry = tk.Entry(form_container, font=('Segoe UI', 12), width=30)
        amount_entry.grid(row=3, column=1, sticky='w', pady=10)

        # Payment Method
        tk.Label(form_container, text="পেমেন্ট:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=4, column=0, sticky='w', pady=10)
        self.cursor.execute('SELECT method_id, method_name FROM payment_methods WHERE is_active = 1')
        methods = self.cursor.fetchall()
        
        payment_var = tk.IntVar()
        payment_frame = tk.Frame(form_container, bg=self.colors['card'])
        payment_frame.grid(row=4, column=1, sticky='w', pady=10)

        if methods:
            payment_var.set(methods[0][0]) # Set default to the first method
            for method_id, method_name in methods:
                tk.Radiobutton(payment_frame, text=method_name, variable=payment_var, value=method_id, bg=self.colors['card']).pack(anchor='w')
        else:
            tk.Label(payment_frame, text="কোন পেমেন্ট পদ্ধতি নেই।", bg=self.colors['card'], fg=self.colors['danger']).pack(anchor='w')
            # Optionally disable save button or prompt user to add payment method


        # Description
        tk.Label(form_container, text="বিবরণ:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=5, column=0, sticky='nw', pady=10)
        desc_text = tk.Text(form_container, font=('Segoe UI', 10), height=3, width=40)
        desc_text.grid(row=5, column=1, sticky='w', pady=10)

        # Buttons
        button_frame = tk.Frame(form_container, bg=self.colors['card'])
        button_frame.grid(row=6, column=0, columnspan=2, pady=20)

        def save():
            try:
                date = date_entry.get()
                account = account_var.get()
                amount_str = amount_entry.get()
                payment = payment_var.get()
                desc = desc_text.get('1.0', 'end-1c')

                if not self.is_numeric(amount_str):
                    messagebox.showerror("ত্রুটি", "পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                amount = float(amount_str)

                if amount <= 0:
                    messagebox.showerror("ত্রুটি", "পরিমাণ অবশ্যই শূন্যের বেশি হতে হবে!")
                    return

                self.cursor.execute('SELECT account_id FROM accounts WHERE account_name = ?', (account,))
                acc_id = self.cursor.fetchone()[0]

                now_time = datetime.now().strftime('%H:%M')

                self.cursor.execute('''
                    INSERT INTO transactions (transaction_date, transaction_time, account_id, amount, payment_method_id, description)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''',
                    (date, now_time, acc_id, amount, payment, desc)
                )
                self.conn.commit()
                messagebox.showinfo("সফল", "লেনদেন সফলভাবে যোগ করা হয়েছে!")
                self.load_dashboard() # Refresh dashboard after saving
            except Exception as e:
                self.logger.error("Failed to add transaction", exc_info=True, extra={'error_message': str(e)})
                messagebox.showerror("ত্রুটি", f"লেনদেন যোগ করতে ব্যর্থ: {e}")

        tk.Button(
            button_frame,
            text="✅ সংরক্ষণ করুন",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['success'],
            fg='white',
            padx=20,
            pady=10,
            bd=0,
            cursor='hand2',
            command=save
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            button_frame,
            text="❌ বাতিল করুন",
            font=('Segoe UI', 12),
            bg=self.colors['danger'],
            fg='white',
            padx=20,
            pady=10,
            bd=0,
            cursor='hand2',
            command=self.show_dashboard
        ).pack(side=tk.LEFT, padx=10)

        # Function to update accounts based on transaction type
        def update_accounts(*args):
            selected_type = trans_type.get()
            self.cursor.execute('SELECT account_name FROM accounts WHERE account_type = ? AND is_active = 1', (selected_type,))
            updated_accounts = [r[0] for r in self.cursor.fetchall()]
            account_combo['values'] = updated_accounts
            if updated_accounts:
                account_combo.current(0)
            else:
                account_var.set("")

        trans_type.trace_add('write', update_accounts)

    def show_reports(self):
        """রিপোর্ট - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="📊 রিপোর্ট",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        report_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        report_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Report Type Selection
        tk.Label(report_frame, text="রিপোর্ট ধরন:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).pack(pady=10)
        report_type_var = tk.StringVar(value="মাসিক")
        report_type_frame = tk.Frame(report_frame, bg=self.colors['card'])
        report_type_frame.pack(pady=5)

        tk.Radiobutton(report_type_frame, text="দৈনিক", variable=report_type_var, value="দৈনিক", bg=self.colors['card']).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(report_type_frame, text="মাসিক", variable=report_type_var, value="মাসিক", bg=self.colors['card']).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(report_type_frame, text="বার্ষিক", variable=report_type_var, value="বার্ষিক", bg=self.colors['card']).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(report_type_frame, text="কাস্টম রেঞ্জ", variable=report_type_var, value="কাস্টম রেঞ্জ", bg=self.colors['card']).pack(side=tk.LEFT, padx=10)

        # Date Range Selection
        date_range_frame = tk.Frame(report_frame, bg=self.colors['card'])
        date_range_frame.pack(pady=10)

        tk.Label(date_range_frame, text="শুরুর তারিখ:", bg=self.colors['card']).pack(side=tk.LEFT)
        start_date_entry = DateEntry(date_range_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd') if CALENDAR_AVAILABLE else tk.Entry(date_range_frame, width=15)
        start_date_entry.pack(side=tk.LEFT, padx=5)
        start_date_entry.set_date(datetime.now().replace(day=1))

        tk.Label(date_range_frame, text="শেষ তারিখ:", bg=self.colors['card']).pack(side=tk.LEFT)
        end_date_entry = DateEntry(date_range_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd') if CALENDAR_AVAILABLE else tk.Entry(date_range_frame, width=15)
        end_date_entry.pack(side=tk.LEFT, padx=5)
        end_date_entry.set_date(datetime.now())

        def update_date_range_visibility():
            if report_type_var.get() == "কাস্টম রেঞ্জ":
                date_range_frame.pack(pady=10)
            else:
                date_range_frame.pack_forget()

        report_type_var.trace_add('write', lambda *args: update_date_range_visibility())
        update_date_range_visibility() # Initial call

        # Report Display Area
        report_display_frame = tk.Frame(report_frame, bg=self.colors['bg'])
        report_display_frame.pack(fill=tk.BOTH, expand=True, pady=10, padx=10)

        report_text = tk.Text(report_display_frame, wrap=tk.WORD, font=('Segoe UI', 10), bg=self.colors['bg'], fg=self.colors['text'])
        report_text.pack(fill=tk.BOTH, expand=True)

        def generate_report():
            report_text.delete('1.0', tk.END)
            r_type = report_type_var.get()
            s_date = start_date_entry.get_date().strftime('%Y-%m-%d') if CALENDAR_AVAILABLE else start_date_entry.get()
            e_date = end_date_entry.get_date().strftime('%Y-%m-%d') if CALENDAR_AVAILABLE else end_date_entry.get()

            if r_type == "দৈনিক":
                s_date = e_date = datetime.now().strftime('%Y-%m-%d')
            elif r_type == "মাসিক":
                s_date = datetime.now().replace(day=1).strftime('%Y-%m-%d')
                e_date = datetime.now().strftime('%Y-%m-%d')
            elif r_type == "বার্ষিক":
                s_date = datetime.now().replace(month=1, day=1).strftime('%Y-%m-%d')
                e_date = datetime.now().strftime('%Y-%m-%d')

            report_text.insert(tk.END, f"রিপোর্ট ধরন: {r_type}\n")
            report_text.insert(tk.END, f"তারিখ রেঞ্জ: {s_date} থেকে {e_date}\n\n")

            self.cursor.execute('''
                SELECT a.account_name, a.account_type, SUM(t.amount)
                FROM transactions t
                JOIN accounts a ON t.account_id = a.account_id
                WHERE t.transaction_date BETWEEN ? AND ? AND t.is_deleted = 0
                GROUP BY a.account_name, a.account_type
                ORDER BY a.account_type DESC, SUM(t.amount) DESC
            ''')
            transactions_summary = self.cursor.fetchall()

            total_income = sum(row[2] for row in transactions_summary if row[1] == 'আয়')
            total_expense = sum(row[2] for row in transactions_summary if row[1] == 'খরচ')

            report_text.insert(tk.END, "আয়:\n")
            for name, type, amount in transactions_summary:
                if type == 'আয়':
                    report_text.insert(tk.END, f"  {name}: ৳ {amount:,.2f}\n")
            report_text.insert(tk.END, f"মোট আয়: ৳ {total_income:,.2f}\n\n")

            report_text.insert(tk.END, "খরচ:\n")
            for name, type, amount in transactions_summary:
                if type == 'খরচ':
                    report_text.insert(tk.END, f"  {name}: ৳ {amount:,.2f}\n")
            report_text.insert(tk.END, f"মোট খরচ: ৳ {total_expense:,.2f}\n\n")

            report_text.insert(tk.END, f"নেট ব্যালেন্স: ৳ {total_income - total_expense:,.2f}\n")

            # Chart
            if CHART_AVAILABLE:
                self.plot_report_chart(s_date, e_date)
            else:
                report_text.insert(tk.END, "\n(Matplotlib ইনস্টল করা নেই, চার্ট দেখানো যাচ্ছে না।)")

        tk.Button(
            report_frame,
            text="জেনারেট রিপোর্ট",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=20,
            pady=10,
            bd=0,
            cursor='hand2',
            command=generate_report
        ).pack(pady=10)

        def export_pdf():
            if not PDF_AVAILABLE:
                messagebox.showerror("ত্রুটি", "ReportLab ইনস্টল করা নেই। পিডিএফ এক্সপোর্ট করা যাবে না।")
                return
            filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
            if not filepath: return

            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            story.append(Paragraph("<b>Advanced Money Tracker - রিপোর্ট</b>", styles['h1']))
            story.append(Spacer(1, 0.2 * inch))
            story.append(Paragraph(f"রিপোর্ট ধরন: {report_type_var.get()}", styles['Normal']))
            story.append(Paragraph(f"তারিখ রেঞ্জ: {start_date_entry.get_date().strftime('%Y-%m-%d')} থেকে {end_date_entry.get_date().strftime('%Y-%m-%d')}", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))

            self.cursor.execute('''
                SELECT a.account_name, a.account_type, SUM(t.amount)
                FROM transactions t
                JOIN accounts a ON t.account_id = a.account_id
                WHERE t.transaction_date BETWEEN ? AND ? AND t.is_deleted = 0
                GROUP BY a.account_name, a.account_type
                ORDER BY a.account_type DESC, SUM(t.amount) DESC
            ''', (start_date_entry.get_date().strftime('%Y-%m-%d'), end_date_entry.get_date().strftime('%Y-%m-%d')))

            data = [["খাত", "ধরন", "পরিমাণ (৳)"]]
            total_income = 0
            total_expense = 0

            for name, type, amount in transactions_summary:
                data.append([name, type, f"{amount:,.2f}"])
                if type == 'আয়':
                    total_income += amount
                else:
                    total_expense += amount

            data.append(["", "", ""])
            data.append(["মোট আয়", "", f"{total_income:,.2f}"])
            data.append(["মোট খরচ", "", f"{total_expense:,.2f}"])
            data.append(["নেট ব্যালেন্স", "", f"{total_income - total_expense:,.2f}"])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), pdf_colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), pdf_colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), pdf_colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, pdf_colors.black)
            ]))
            story.append(table)

            try:
                doc.build(story)
                messagebox.showinfo("সফল", "পিডিএফ রিপোর্ট সফলভাবে এক্সপোর্ট করা হয়েছে!")
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"পিডিএফ এক্সপোর্ট করতে ব্যর্থ: {e}")

        tk.Button(
            report_frame,
            text="পিডিএফ এক্সপোর্ট",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=20,
            pady=10,
            bd=0,
            cursor='hand2',
            command=export_pdf
        ).pack(pady=10)

    def plot_report_chart(self, start_date, end_date):
        for widget in self.content_area.winfo_children():
            if isinstance(widget, tk.Frame) and "chart_frame" in str(widget):
                widget.destroy()

        chart_frame = tk.Frame(self.content_area, bg=self.colors['card'], name="chart_frame")
        chart_frame.pack(fill=tk.BOTH, expand=True, pady=10, padx=10)

        self.cursor.execute('''
            SELECT a.account_type, SUM(t.amount)
            FROM transactions t
            JOIN accounts a ON t.account_id = a.account_id
            WHERE t.transaction_date BETWEEN ? AND ? AND t.is_deleted = 0
            GROUP BY a.account_type
        ''')
        data = self.cursor.fetchall()

        if not data:
            tk.Label(chart_frame, text="এই সময়ের জন্য কোন ডেটা নেই।", bg=self.colors['card'], fg=self.colors['text']).pack(pady=20)
            return

        labels = [row[0] for row in data]
        sizes = [row[1] for row in data]
        colors = [self.colors['success'] if label == 'আয়' else self.colors['danger'] for label in labels]

        fig = Figure(figsize=(6, 4), dpi=100, facecolor=self.colors['card'])
        ax = fig.add_subplot(111)
        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        ax.set_title('আয় বনাম খরচ', color=self.colors['text'])

        canvas = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        canvas.draw()

    def show_budget(self):
        """বাজেট - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="💰 বাজেট",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        budget_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        budget_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Budget Entry Form
        form_frame = tk.Frame(budget_frame, bg=self.colors['card'])
        form_frame.pack(pady=10)

        tk.Label(form_frame, text="খাত:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.cursor.execute('SELECT account_name FROM accounts WHERE account_type = "খরচ" AND is_active = 1')
        accounts = [r[0] for r in self.cursor.fetchall()]
        budget_account_var = tk.StringVar()
        budget_account_combo = ttk.Combobox(form_frame, textvariable=budget_account_var, values=accounts, width=30, state='readonly')
        budget_account_combo.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        if accounts: budget_account_combo.current(0)

        tk.Label(form_frame, text="মাস:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        months = [datetime(2000, i, 1).strftime('%Y-%m') for i in range(1, 13)] # YYYY-MM format
        budget_month_var = tk.StringVar(value=datetime.now().strftime('%Y-%m'))
        budget_month_combo = ttk.Combobox(form_frame, textvariable=budget_month_var, values=months, width=30, state='readonly')
        budget_month_combo.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="বাজেট পরিমাণ (৳):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        budget_amount_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        budget_amount_entry.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        def is_numeric(value):
            try:
                float(value)
                return True
            except ValueError:
                return False

        def save_budget():
            try:
                account_name = budget_account_var.get()
                month = budget_month_var.get()
                amount_str = budget_amount_entry.get()

                if not is_numeric(amount_str):
                    messagebox.showerror("ত্রুটি", "বাজেট পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                amount = float(amount_str)

                if amount <= 0:
                    messagebox.showerror("ত্রুটি", "বাজেট পরিমাণ অবশ্যই শূন্যের বেশি হতে হবে!")
                    return

                self.cursor.execute('SELECT account_id FROM accounts WHERE account_name = ?', (account_name,))
                account_id = self.cursor.fetchone()[0]

                self.cursor.execute('''
                    INSERT OR REPLACE INTO budget (account_id, month, budgeted_amount)
                    VALUES (?, ?, ?)
                ''', (account_id, month, amount))
                self.conn.commit()
                messagebox.showinfo("সফল", "বাজেট সফলভাবে সংরক্ষণ করা হয়েছে!")
                self.load_budget_list()
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"বাজেট সংরক্ষণ করতে ব্যর্থ: {e}")

        tk.Button(
            form_frame,
            text="সংরক্ষণ বাজেট",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=save_budget
        ).grid(row=3, column=0, columnspan=2, pady=10)

        # Budget List
        budget_list_frame = tk.Frame(budget_frame, bg=self.colors['card'])
        budget_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('মাস', 'খাত', 'বাজেট পরিমাণ', 'খরচ', 'অবশিষ্ট', 'স্ট্যাটাস')
        self.budget_tree = ttk.Treeview(budget_list_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.budget_tree.heading(col, text=col)
            self.budget_tree.column(col, width=100)

        self.budget_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(budget_list_frame, orient=tk.VERTICAL, command=self.budget_tree.yview)
        self.budget_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.load_budget_list()

    def load_budget_list(self):
        for item in self.budget_tree.get_children():
            self.budget_tree.delete(item)

        self.cursor.execute('''
            SELECT b.month, a.account_name, b.budgeted_amount, a.account_id
            FROM budget b
            JOIN accounts a ON b.account_id = a.account_id
            ORDER BY b.month DESC, a.account_name ASC
        ''')
        budgets = self.cursor.fetchall()

        # Fetch all actual expenses in a single query
        self.cursor.execute('''
            SELECT STRFTIME('%Y-%m', t.transaction_date) AS month, t.account_id, COALESCE(SUM(t.amount), 0) AS total_expense
            FROM transactions t
            WHERE t.is_deleted = 0
            GROUP BY month, t.account_id
        ''')
        actual_expenses_data = self.cursor.fetchall()
        
        # Create a dictionary for quick lookup: {(month, account_id): total_expense}
        actual_expenses_map = {}
        for month, account_id, total_expense in actual_expenses_data:
            actual_expenses_map[(month, account_id)] = total_expense

        for month, account_name, budgeted_amount, account_id in budgets:
            actual_expense = actual_expenses_map.get((month, account_id), 0)

            remaining = budgeted_amount - actual_expense
            status = "ঠিক আছে" if remaining >= 0 else "বেশি খরচ"

            self.budget_tree.insert('', 'end', values=(month, account_name, f"৳ {budgeted_amount:,.2f}", f"৳ {actual_expense:,.2f}", f"৳ {remaining:,.2f}", status))

    def show_recurring(self):
        """Recurring Transactions - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="🔄 Recurring লেনদেন",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        recurring_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        recurring_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Form for new recurring transaction
        form_frame = tk.Frame(recurring_frame, bg=self.colors['card'])
        form_frame.pack(pady=10)

        tk.Label(form_frame, text="খাত:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.cursor.execute('SELECT account_name FROM accounts WHERE is_active = 1')
        accounts = [r[0] for r in self.cursor.fetchall()]
        rec_account_var = tk.StringVar()
        rec_account_combo = ttk.Combobox(form_frame, textvariable=rec_account_var, values=accounts, width=30, state='readonly')
        rec_account_combo.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        if accounts: rec_account_combo.current(0)

        tk.Label(form_frame, text="পরিমাণ (৳):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        rec_amount_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        rec_amount_entry.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="বিবরণ:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        rec_desc_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        rec_desc_entry.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="ফ্রিকোয়েন্সি:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=3, column=0, sticky='w', padx=5, pady=5)
        frequencies = ['দৈনিক', 'সাপ্তাহিক', 'মাসিক', 'বার্ষিক']
        rec_frequency_var = tk.StringVar(value='মাসিক')
        rec_frequency_combo = ttk.Combobox(form_frame, textvariable=rec_frequency_var, values=frequencies, width=30, state='readonly')
        rec_frequency_combo.grid(row=3, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="মাসের দিন (মাসিকের জন্য):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=4, column=0, sticky='w', padx=5, pady=5)
        rec_day_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        rec_day_entry.grid(row=4, column=1, sticky='w', padx=5, pady=5)
        rec_day_entry.insert(0, "1")

        def save_recurring():
            try:
                account_name = rec_account_var.get()
                amount_str = rec_amount_entry.get()
                description = rec_desc_entry.get()
                frequency = rec_frequency_var.get()
                day_of_month_str = rec_day_entry.get()

                if not self.is_numeric(amount_str):
                    messagebox.showerror("ত্রুটি", "পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                amount = float(amount_str)

                if amount <= 0:
                    messagebox.showerror("ত্রুটি", "পরিমাণ অবশ্যই শূন্যের বেশি হতে হবে!")
                    return

                day_of_month = None
                if frequency == 'মাসিক':
                    if not day_of_month_str.isdigit():
                        messagebox.showerror("ত্রুটি", "মাসিক ফ্রিকোয়েন্সির জন্য মাসের দিন অবশ্যই একটি পূর্ণসংখ্যা হতে হবে!")
                        return
                    day_of_month = int(day_of_month_str)
                    if not (1 <= day_of_month <= 31):
                        messagebox.showerror("ত্রুটি", "মাসের দিন 1 থেকে 31 এর মধ্যে হতে হবে!")
                        return

                self.cursor.execute('SELECT account_id FROM accounts WHERE account_name = ?', (account_name,))
                acc_id = self.cursor.fetchone()[0]

                # Determine next_date
                next_date = datetime.now()
                if frequency == 'দৈনিক':
                    pass # next_date is today
                elif frequency == 'সাপ্তাহিক':
                    next_date += timedelta(days=(7 - next_date.weekday()) % 7) # Next Sunday
                elif frequency == 'মাসিক':
                    if day_of_month is not None:
                        try:
                            next_date = next_date.replace(day=day_of_month)
                            if next_date < datetime.now():
                                next_date = next_date.replace(month=next_date.month % 12 + 1)
                        except ValueError: # Day out of range for month
                            messagebox.showerror("ত্রুটি", "এই মাসের জন্য দিনের সংখ্যা সঠিক নয়।")
                            return
                    else:
                        messagebox.showerror("ত্রুটি", "মাসিক ফ্রিকোয়েন্সির জন্য মাসের দিন প্রয়োজন।")
                        return
                elif frequency == 'বার্ষিক':
                    next_date = next_date.replace(year=next_date.year + 1)

                self.cursor.execute('''
                    INSERT INTO recurring_templates (account_id, amount, description, frequency, day_of_month, next_date, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                ''', (acc_id, amount, description, frequency, day_of_month, next_date.strftime('%Y-%m-%d')))
                self.conn.commit()
                messagebox.showinfo("সফল", "Recurring লেনদেন সফলভাবে যোগ করা হয়েছে!")
                self.load_recurring_list()
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"Recurring লেনদেন যোগ করতে ব্যর্থ: {e}")

        tk.Button(
            form_frame,
            text="সংরক্ষণ Recurring",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=save_recurring
        ).grid(row=5, column=0, columnspan=2, pady=10)

        # Recurring List
        recurring_list_frame = tk.Frame(recurring_frame, bg=self.colors['card'])
        recurring_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('খাত', 'পরিমাণ', 'ফ্রিকোয়েন্সি', 'পরবর্তী তারিখ', 'স্ট্যাটাস')
        self.recurring_tree = ttk.Treeview(recurring_list_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.recurring_tree.heading(col, text=col)
            self.recurring_tree.column(col, width=120)

        self.recurring_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(recurring_list_frame, orient=tk.VERTICAL, command=self.recurring_tree.yview)
        self.recurring_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.load_recurring_list()

    def load_recurring_list(self):
        for item in self.recurring_tree.get_children():
            self.recurring_tree.delete(item)

        self.cursor.execute('''
            SELECT rt.template_id, a.account_name, rt.amount, rt.frequency, rt.next_date, rt.is_active
            FROM recurring_templates rt
            JOIN accounts a ON rt.account_id = a.account_id
            ORDER BY rt.next_date ASC
        ''')
        recurring_items = self.cursor.fetchall()

        for template_id, account_name, amount, frequency, next_date, is_active in recurring_items:
            status = "সক্রিয়" if is_active else "নিষ্ক্রিয়"
            self.recurring_tree.insert('', 'end', values=(account_name, f"৳ {amount:,.2f}", frequency, next_date, status), iid=template_id)

        self.recurring_tree.bind('<Double-1>', self.edit_recurring_item)

    def edit_recurring_item(self, event):
        selected_item = self.recurring_tree.selection()
        if not selected_item: return

        item_id = self.recurring_tree.focus()
        values = self.recurring_tree.item(item_id, 'values')

        # Fetch full details from DB
        self.cursor.execute('SELECT * FROM recurring_templates WHERE template_id = ?', (item_id,))
        template_data = self.cursor.fetchone()
        if not template_data: return

        # Create a new window for editing
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Recurring লেনদেন সম্পাদনা")
        edit_window.geometry("400x400")
        edit_window.transient(self.root)
        edit_window.grab_set()

        labels = ["খাত:", "পরিমাণ (৳):", "বিবরণ:", "ফ্রিকোয়েন্সি:", "মাসের দিন:", "সক্রিয়?"]
        entries = {}

        # Account
        tk.Label(edit_window, text=labels[0]).grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.cursor.execute('SELECT account_name FROM accounts WHERE account_type = "খরচ" AND is_active = 1')
        accounts = [r[0] for r in self.cursor.fetchall()]
        acc_var = tk.StringVar(value=values[0])
        acc_combo = ttk.Combobox(edit_window, textvariable=acc_var, values=accounts, width=30, state='readonly')
        acc_combo.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        entries['account_id'] = acc_combo

        # Amount
        tk.Label(edit_window, text=labels[1]).grid(row=1, column=0, padx=5, pady=5, sticky='w')
        amount_entry = tk.Entry(edit_window, width=30)
        amount_entry.insert(0, template_data[3])
        amount_entry.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        entries['amount'] = amount_entry

        # Description
        tk.Label(edit_window, text=labels[2]).grid(row=2, column=0, padx=5, pady=5, sticky='w')
        desc_entry = tk.Entry(edit_window, width=30)
        desc_entry.insert(0, template_data[4])
        desc_entry.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        entries['description'] = desc_entry

        # Frequency
        tk.Label(edit_window, text=labels[3]).grid(row=3, column=0, padx=5, pady=5, sticky='w')
        frequencies = ['দৈনিক', 'সাপ্তাহিক', 'মাসিক', 'বার্ষিক']
        freq_var = tk.StringVar(value=template_data[5])
        freq_combo = ttk.Combobox(edit_window, textvariable=freq_var, values=frequencies, width=30, state='readonly')
        freq_combo.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        entries['frequency'] = freq_combo

        # Day of Month
        tk.Label(edit_window, text=labels[4]).grid(row=4, column=0, padx=5, pady=5, sticky='w')
        day_entry = tk.Entry(edit_window, width=30)
        day_entry.insert(0, template_data[6] if template_data[6] else "")
        day_entry.grid(row=4, column=1, padx=5, pady=5, sticky='ew')
        entries['day_of_month'] = day_entry

        # Is Active
        tk.Label(edit_window, text=labels[5]).grid(row=5, column=0, padx=5, pady=5, sticky='w')
        is_active_var = tk.IntVar(value=template_data[7])
        tk.Checkbutton(edit_window, variable=is_active_var).grid(row=5, column=1, padx=5, pady=5, sticky='w')
        entries['is_active'] = is_active_var

        def update_recurring():
            try:
                account_name = acc_var.get()
                amount_str = amount_entry.get()
                description = desc_entry.get()
                frequency = freq_var.get()
                day_of_month_str = day_entry.get()
                is_active = is_active_var.get()

                if not self.is_numeric(amount_str):
                    messagebox.showerror("ত্রুটি", "পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                amount = float(amount_str)

                if amount <= 0:
                    messagebox.showerror("ত্রুটি", "পরিমাণ অবশ্যই শূন্যের বেশি হতে হবে!")
                    return

                day_of_month = None
                if frequency == 'মাসিক':
                    if not day_of_month_str.isdigit():
                        messagebox.showerror("ত্রুটি", "মাসিক ফ্রিকোয়েন্সির জন্য মাসের দিন অবশ্যই একটি পূর্ণসংখ্যা হতে হবে!")
                        return
                    day_of_month = int(day_of_month_str)
                    if not (1 <= day_of_month <= 31):
                        messagebox.showerror("ত্রুটি", "মাসের দিন 1 থেকে 31 এর মধ্যে হতে হবে!")
                        return

                self.cursor.execute('SELECT account_id FROM accounts WHERE account_name = ?', (account_name,))
                account_id = self.cursor.fetchone()[0]

                # Recalculate next_date if frequency or day_of_month changed
                current_next_date = datetime.strptime(template_data[8], '%Y-%m-%d') if template_data[8] else datetime.now()
                new_next_date = current_next_date

                if frequency == 'দৈনিক':
                    new_next_date = datetime.now() # Or keep current if it's in the future
                elif frequency == 'সাপ্তাহিক':
                    new_next_date += timedelta(days=(7 - new_next_date.weekday()) % 7)
                elif frequency == 'মাসিক':
                    if day_of_month is not None:
                        try:
                            new_next_date = new_next_date.replace(day=day_of_month)
                            if new_next_date < datetime.now():
                                new_next_date = new_next_date.replace(month=new_next_date.month % 12 + 1)
                        except ValueError: # Day out of range for month
                            messagebox.showerror("ত্রুটি", "এই মাসের জন্য দিনের সংখ্যা সঠিক নয়।")
                            return
                    else:
                        messagebox.showerror("ত্রুটি", "মাসিক ফ্রিকোয়েন্সির জন্য মাসের দিন প্রয়োজন।")
                        return
                elif frequency == 'বার্ষিক':
                    new_next_date = new_next_date.replace(year=new_next_date.year + 1)

                self.cursor.execute('''
                    UPDATE recurring_templates
                    SET account_id = ?, amount = ?, description = ?, frequency = ?, day_of_month = ?, is_active = ?, next_date = ?
                    WHERE template_id = ?
                ''',
                    (account_id, amount, description, frequency, day_of_month, is_active, new_next_date.strftime('%Y-%m-%d'), item_id)
                )
                self.conn.commit()
                messagebox.showinfo("সফল", "Recurring লেনদেন সফলভাবে আপডেট করা হয়েছে!")
                edit_window.destroy()
                self.load_recurring_list()
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"Recurring লেনদেন আপডেট করতে ব্যর্থ: {e}")

        tk.Button(edit_window, text="আপডেট", command=update_recurring).grid(row=6, column=0, padx=5, pady=10)
        tk.Button(edit_window, text="বাতিল", command=edit_window.destroy).grid(row=6, column=1, padx=5, pady=10)

    def show_goals(self):
        """Financial Goals - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="🎯 আর্থিক লক্ষ্য",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        goals_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        goals_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Form for new goal
        form_frame = tk.Frame(goals_frame, bg=self.colors['card'])
        form_frame.pack(pady=10)

        tk.Label(form_frame, text="লক্ষ্যের নাম:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        goal_name_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        goal_name_entry.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="লক্ষ্য পরিমাণ (৳):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        target_amount_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        target_amount_entry.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="সময়সীমা (YYYY-MM-DD):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        deadline_entry = DateEntry(form_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd') if CALENDAR_AVAILABLE else tk.Entry(form_frame, width=30)
        deadline_entry.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        deadline_entry.set_date(datetime.now() + timedelta(days=365))

        tk.Label(form_frame, text="ক্যাটাগরি:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=3, column=0, sticky='w', padx=5, pady=5)
        goal_category_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        goal_category_entry.grid(row=3, column=1, sticky='w', padx=5, pady=5)

        def save_goal():
            try:
                goal_name = goal_name_entry.get()
                target_amount_str = target_amount_entry.get()
                deadline = deadline_entry.get_date().strftime('%Y-%m-%d') if CALENDAR_AVAILABLE else deadline_entry.get()
                category = goal_category_entry.get()

                if not self.is_numeric(target_amount_str):
                    messagebox.showerror("ত্রুটি", "লক্ষ্য পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                target_amount = float(target_amount_str)

                if target_amount <= 0:
                    messagebox.showerror("ত্রুটি", "লক্ষ্য পরিমাণ অবশ্যই শূন্যের বেশি হতে হবে!")
                    return

                self.cursor.execute('''
                    INSERT INTO financial_goals (goal_name, target_amount, deadline, category)
                    VALUES (?, ?, ?, ?)
                ''', (goal_name, target_amount, deadline, category))
                self.conn.commit()
                messagebox.showinfo("সফল", "আর্থিক লক্ষ্য সফলভাবে যোগ করা হয়েছে!")
                self.load_goals_list()
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"আর্থিক লক্ষ্য যোগ করতে ব্যর্থ: {e}")

        tk.Button(
            form_frame,
            text="সংরক্ষণ লক্ষ্য",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=save_goal
        ).grid(row=4, column=0, columnspan=2, pady=10)

        # Goals List
        goals_list_frame = tk.Frame(goals_frame, bg=self.colors['card'])
        goals_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('লক্ষ্যের নাম', 'লক্ষ্য পরিমাণ', 'বর্তমান পরিমাণ', 'সময়সীমা', 'স্ট্যাটাস')
        self.goals_tree = ttk.Treeview(goals_list_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.goals_tree.heading(col, text=col)
            self.goals_tree.column(col, width=120)

        self.goals_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(goals_list_frame, orient=tk.VERTICAL, command=self.goals_tree.yview)
        self.goals_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.load_goals_list()

    def load_goals_list(self):
        for item in self.goals_tree.get_children():
            self.goals_tree.delete(item)

        self.cursor.execute('SELECT goal_id, goal_name, target_amount, current_amount, deadline, is_achieved FROM financial_goals ORDER BY deadline ASC')
        goals = self.cursor.fetchall()

        for goal_id, goal_name, target_amount, current_amount, deadline, is_achieved in goals:
            status = "সম্পন্ন" if is_achieved else "চলমান"
            progress = (current_amount / target_amount) * 100 if target_amount > 0 else 0
            display_status = f"{status} ({progress:.2f}%)"
            self.goals_tree.insert('', 'end', values=(goal_name, f"৳ {target_amount:,.2f}", f"৳ {current_amount:,.2f}", deadline, display_status), iid=goal_id)

        self.goals_tree.bind('<Double-1>', self.edit_goal_item)

    def edit_goal_item(self, event):
        selected_item = self.goals_tree.selection()
        if not selected_item: return

        item_id = self.goals_tree.focus()
        self.cursor.execute('SELECT * FROM financial_goals WHERE goal_id = ?', (item_id,))
        goal_data = self.cursor.fetchone()
        if not goal_data: return

        edit_window = tk.Toplevel(self.root)
        edit_window.title("আর্থিক লক্ষ্য সম্পাদনা")
        edit_window.geometry("400x350")
        edit_window.transient(self.root)
        edit_window.grab_set()

        labels = ["লক্ষ্যের নাম:", "লক্ষ্য পরিমাণ (৳):", "বর্তমান পরিমাণ (৳):", "সময়সীমা (YYYY-MM-DD):", "ক্যাটাগরি:", "সম্পন্ন?"]
        entries = {}

        for i, label_text in enumerate(labels):
            tk.Label(edit_window, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky='w')

        goal_name_entry = tk.Entry(edit_window, width=30)
        goal_name_entry.insert(0, goal_data[1])
        goal_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        entries['goal_name'] = goal_name_entry

        target_amount_entry = tk.Entry(edit_window, width=30)
        target_amount_entry.insert(0, goal_data[2])
        target_amount_entry.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        entries['target_amount'] = target_amount_entry

        current_amount_entry = tk.Entry(edit_window, width=30)
        current_amount_entry.insert(0, goal_data[3])
        current_amount_entry.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        entries['current_amount'] = current_amount_entry

        deadline_entry = DateEntry(edit_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd') if CALENDAR_AVAILABLE else tk.Entry(edit_window, width=30)
        deadline_entry.set_date(goal_data[4])
        deadline_entry.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        entries['deadline'] = deadline_entry

        category_entry = tk.Entry(edit_window, width=30)
        category_entry.insert(0, goal_data[5])
        category_entry.grid(row=4, column=1, padx=5, pady=5, sticky='ew')
        entries['category'] = category_entry

        is_achieved_var = tk.IntVar(value=goal_data[7])
        tk.Checkbutton(edit_window, variable=is_achieved_var).grid(row=5, column=1, padx=5, pady=5, sticky='w')
        entries['is_achieved'] = is_achieved_var

        def update_goal():
            try:
                goal_name = goal_name_entry.get()
                target_amount_str = target_amount_entry.get()
                current_amount_str = current_amount_entry.get()
                deadline = deadline_entry.get_date().strftime('%Y-%m-%d') if CALENDAR_AVAILABLE else deadline_entry.get()
                category = category_entry.get()
                is_achieved = is_achieved_var.get()

                if not self.is_numeric(target_amount_str):
                    messagebox.showerror("ত্রুটি", "লক্ষ্য পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                target_amount = float(target_amount_str)

                if target_amount < 0:
                    messagebox.showerror("ত্রুটি", "লক্ষ্য পরিমাণ অবশ্যই শূন্য বা তার বেশি হতে হবে!")
                    return

                if not self.is_numeric(current_amount_str):
                    messagebox.showerror("ত্রুটি", "বর্তমান পরিমাণ অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                current_amount = float(current_amount_str)

                if current_amount < 0:
                    messagebox.showerror("ত্রুটি", "বর্তমান পরিমাণ অবশ্যই শূন্য বা তার বেশি হতে হবে!")
                    return

                self.cursor.execute('''
                    UPDATE financial_goals
                    SET goal_name = ?, target_amount = ?, current_amount = ?, deadline = ?, category = ?, is_achieved = ?
                    WHERE goal_id = ?
                ''',
                    (goal_name, target_amount, current_amount, deadline, category, is_achieved, item_id)
                )
                self.conn.commit()
                messagebox.showinfo("সফল", "আর্থিক লক্ষ্য সফলভাবে আপডেট করা হয়েছে!")
                edit_window.destroy()
                self.load_goals_list()
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"আর্থিক লক্ষ্য আপডেট করতে ব্যর্থ: {e}")

        tk.Button(edit_window, text="আপডেট", command=update_goal).grid(row=6, column=0, padx=5, pady=10)
        tk.Button(edit_window, text="বাতিল", command=edit_window.destroy).grid(row=6, column=1, padx=5, pady=10)

    def show_payment_methods(self):
        """Payment Methods - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="💳 পেমেন্ট পদ্ধতি",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        pm_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        pm_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Form for new payment method
        form_frame = tk.Frame(pm_frame, bg=self.colors['card'])
        form_frame.pack(pady=10)

        tk.Label(form_frame, text="পদ্ধতির নাম:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        method_name_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        method_name_entry.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="ধরন:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        method_types = ['নগদ', 'ব্যাংক', 'মোবাইল ব্যাংকিং', 'কার্ড']
        method_type_var = tk.StringVar(value='নগদ')
        method_type_combo = ttk.Combobox(form_frame, textvariable=method_type_var, values=method_types, width=30, state='readonly')
        method_type_combo.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="প্রাথমিক ব্যালেন্স (৳):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        initial_balance_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        initial_balance_entry.insert(0, "0.00")
        initial_balance_entry.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        def save_method():
            try:
                method_name = method_name_entry.get()
                method_type = method_type_var.get()
                initial_balance_str = initial_balance_entry.get()

                if not self.is_numeric(initial_balance_str):
                    messagebox.showerror("ত্রুটি", "প্রাথমিক ব্যালেন্স অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                initial_balance = float(initial_balance_str)

                if initial_balance < 0:
                    messagebox.showerror("ত্রুটি", "প্রাথমিক ব্যালেন্স অবশ্যই শূন্য বা তার বেশি হতে হবে!")
                    return

                self.cursor.execute('''
                    INSERT INTO payment_methods (method_name, method_type, balance)
                    VALUES (?, ?, ?)
                ''', (method_name, method_type, initial_balance))
                self.conn.commit()
                messagebox.showinfo("সফল", "পেমেন্ট পদ্ধতি সফলভাবে যোগ করা হয়েছে!")
                self.load_payment_methods_list()
            except sqlite3.IntegrityError:
                messagebox.showerror("ত্রুটি", "এই নামের একটি পেমেন্ট পদ্ধতি ইতিমধ্যেই বিদ্যমান।")
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"পেমেন্ট পদ্ধতি যোগ করতে ব্যর্থ: {e}")

        tk.Button(
            form_frame,
            text="সংরক্ষণ পদ্ধতি",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=save_method
        ).grid(row=3, column=0, columnspan=2, pady=10)

        # Payment Methods List
        pm_list_frame = tk.Frame(pm_frame, bg=self.colors['card'])
        pm_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('পদ্ধতির নাম', 'ধরন', 'ব্যালেন্স', 'সক্রিয়?')
        self.pm_tree = ttk.Treeview(pm_list_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.pm_tree.heading(col, text=col)
            self.pm_tree.column(col, width=120)

        self.pm_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(pm_list_frame, orient=tk.VERTICAL, command=self.pm_tree.yview)
        self.pm_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.load_payment_methods_list()

    def load_payment_methods_list(self):
        for item in self.pm_tree.get_children():
            self.pm_tree.delete(item)

        self.cursor.execute('SELECT method_id, method_name, method_type, balance, is_active FROM payment_methods ORDER BY method_name ASC')
        methods = self.cursor.fetchall()

        for method_id, method_name, method_type, balance, is_active in methods:
            status = "সক্রিয়" if is_active else "নিষ্ক্রিয়"
            self.pm_tree.insert('', 'end', values=(method_name, method_type, f"৳ {balance:,.2f}", status), iid=method_id)

        self.pm_tree.bind('<Double-1>', self.edit_payment_method)

    def edit_payment_method(self, event):
        selected_item = self.pm_tree.selection()
        if not selected_item: return

        item_id = self.pm_tree.focus()
        self.cursor.execute('SELECT * FROM payment_methods WHERE method_id = ?', (item_id,))
        method_data = self.cursor.fetchone()
        if not method_data: return

        edit_window = tk.Toplevel(self.root)
        edit_window.title("পেমেন্ট পদ্ধতি সম্পাদনা")
        edit_window.geometry("400x300")
        edit_window.transient(self.root)
        edit_window.grab_set()

        labels = ["পদ্ধতির নাম:", "ধরন:", "ব্যালেন্স (৳):", "সক্রিয়?"]
        entries = {}

        for i, label_text in enumerate(labels):
            tk.Label(edit_window, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky='w')

        method_name_entry = tk.Entry(edit_window, width=30)
        method_name_entry.insert(0, method_data[1])
        method_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        entries['method_name'] = method_name_entry

        method_types = ['নগদ', 'ব্যাংক', 'মোবাইল ব্যাংকিং', 'কার্ড']
        method_type_var = tk.StringVar(value=method_data[2])
        method_type_combo = ttk.Combobox(edit_window, textvariable=method_type_var, values=method_types, width=30, state='readonly')
        method_type_combo.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        entries['method_type'] = method_type_combo

        balance_entry = tk.Entry(edit_window, width=30)
        balance_entry.insert(0, method_data[3])
        balance_entry.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        entries['balance'] = balance_entry

        is_active_var = tk.IntVar(value=method_data[5])
        tk.Checkbutton(edit_window, variable=is_active_var).grid(row=3, column=1, padx=5, pady=5, sticky='w')
        entries['is_active'] = is_active_var

        def update_method():
            try:
                method_name = method_name_entry.get()
                method_type = method_type_var.get()
                balance_str = balance_entry.get()
                is_active = is_active_var.get()

                if not self.is_numeric(balance_str):
                    messagebox.showerror("ত্রুটি", "ব্যালেন্স অবশ্যই একটি সংখ্যা হতে হবে!")
                    return
                balance = float(balance_str)

                if balance < 0:
                    messagebox.showerror("ত্রুটি", "ব্যালেন্স অবশ্যই শূন্য বা তার বেশি হতে হবে!")
                    return

                self.cursor.execute('''
                    UPDATE payment_methods
                    SET method_name = ?, method_type = ?, balance = ?, is_active = ?
                    WHERE method_id = ?
                ''',
                    (method_name, method_type, balance, is_active, item_id)
                )
                self.conn.commit()
                messagebox.showinfo("সফল", "পেমেন্ট পদ্ধতি সফলভাবে আপডেট করা হয়েছে!")
                edit_window.destroy()
                self.load_payment_methods_list()
            except sqlite3.IntegrityError:
                messagebox.showerror("ত্রুটি", "এই নামের একটি পেমেন্ট পদ্ধতি ইতিমধ্যেই বিদ্যমান।")
            except Exception as e:
                messagebox.showerror("ত্রুটি", f"পেমেন্ট পদ্ধতি আপডেট করতে ব্যর্থ: {e}")

        tk.Button(edit_window, text="আপডেট", command=update_method).grid(row=4, column=0, padx=5, pady=10)
        tk.Button(edit_window, text="বাতিল", command=edit_window.destroy).grid(row=4, column=1, padx=5, pady=10)

    def show_export(self):
        """Export Data - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="📤 ডেটা এক্সপোর্ট",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        export_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        export_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        tk.Label(export_frame, text="এক্সপোর্ট ফরম্যাট:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).pack(pady=10)
        export_format_var = tk.StringVar(value="Excel")
        tk.Radiobutton(export_frame, text="Excel (.xlsx)", variable=export_format_var, value="Excel", bg=self.colors['card']).pack()
        tk.Radiobutton(export_frame, text="PDF (.pdf)", variable=export_format_var, value="PDF", bg=self.colors['card']).pack()

        tk.Label(export_frame, text="শুরুর তারিখ:", bg=self.colors['card']).pack(pady=5)
        start_date_entry = DateEntry(export_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd') if CALENDAR_AVAILABLE else tk.Entry(export_frame, width=15)
        start_date_entry.pack(pady=2)
        start_date_entry.set_date(datetime.now().replace(day=1))

        tk.Label(export_frame, text="শেষ তারিখ:", bg=self.colors['card']).pack(pady=5)
        end_date_entry = DateEntry(export_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd') if CALENDAR_AVAILABLE else tk.Entry(export_frame, width=15)
        end_date_entry.pack(pady=2)
        end_date_entry.set_date(datetime.now())

        def perform_export():
            export_format = export_format_var.get()
            start_date = start_date_entry.get_date().strftime('%Y-%m-%d') if CALENDAR_AVAILABLE else start_date_entry.get()
            end_date = end_date_entry.get_date().strftime('%Y-%m-%d') if CALENDAR_AVAILABLE else end_date_entry.get()

            self.cursor.execute('''
                SELECT t.transaction_date, t.transaction_time, a.account_name, a.account_type, t.amount, t.description, pm.method_name
                FROM transactions t
                JOIN accounts a ON t.account_id = a.account_id
                LEFT JOIN payment_methods pm ON t.payment_method_id = pm.method_id
                WHERE t.transaction_date BETWEEN ? AND ? AND t.is_deleted = 0
                ORDER BY t.transaction_date ASC, t.transaction_time ASC
            ''', (start_date, end_date))
            data_to_export = self.cursor.fetchall()

            headers = ["তারিখ", "সময়", "খাত", "ধরন", "পরিমাণ", "বিবরণ", "পেমেন্ট পদ্ধতি"]

            if export_format == "Excel":
                if not EXCEL_AVAILABLE:
                    messagebox.showerror("ত্রুটি", "openpyxl ইনস্টল করা নেই। এক্সেল এক্সপোর্ট করা যাবে না।")
                    return
                filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if not filepath: return
                self.export_to_excel(filepath, headers, data_to_export)
            elif export_format == "PDF":
                if not PDF_AVAILABLE:
                    messagebox.showerror("ত্রুটি", "ReportLab ইনস্টল করা নেই। পিডিএফ এক্সপোর্ট করা যাবে না।")
                    return
                filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
                if not filepath: return
                self.export_to_pdf(filepath, headers, data_to_export, start_date, end_date)

        tk.Button(
            export_frame,
            text="এক্সপোর্ট করুন",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=20,
            pady=10,
            bd=0,
            cursor='hand2',
            command=perform_export
        ).pack(pady=20)

    def export_to_excel(self, filepath, headers, data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Write headers
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).font = Font(bold=True)
            ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=col_num).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write data
        for row_data in data:
            ws.append(row_data)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        try:
            wb.save(filepath)
            messagebox.showinfo("সফল", f"ডেটা সফলভাবে এক্সেল-এ এক্সপোর্ট করা হয়েছে:\n{filepath}")
        except Exception as e:
            messagebox.showerror("ত্রুটি", f"এক্সেল এক্সপোর্ট করতে ব্যর্থ: {e}")

    def export_to_pdf(self, filepath, headers, data, start_date, end_date):
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("<b>Advanced Money Tracker - লেনদেন রিপোর্ট</b>", styles['h1']))
        story.append(Spacer(1, 0.2 * inch))
        story.append(Paragraph(f"তারিখ রেঞ্জ: {start_date} থেকে {end_date}", styles['Normal']))
        story.append(Spacer(1, 0.2 * inch))

        table_data = [headers] + list(data)
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), pdf_colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), pdf_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), pdf_colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, pdf_colors.black)
        ]))
        story.append(table)

        try:
            doc.build(story)
            messagebox.showinfo("সফল", f"ডেটা সফলভাবে পিডিএফ-এ এক্সপোর্ট করা হয়েছে:\n{filepath}")
        except Exception as e:
            messagebox.showerror("ত্রুটি", f"পিডিএফ এক্সপোর্ট করতে ব্যর্থ: {e}")

    def show_search(self):
        """Search Transactions - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="🔍 লেনদেন খুঁজুন",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        search_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        search_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Search Form
        form_frame = tk.Frame(search_frame, bg=self.colors['card'])
        form_frame.pack(pady=10)

        tk.Label(form_frame, text="অনুসন্ধান শব্দ:", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        search_term_entry = tk.Entry(form_frame, font=('Segoe UI', 11), width=30)
        search_term_entry.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        tk.Label(form_frame, text="খাত (ঐচ্ছিক):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.cursor.execute('SELECT account_name FROM accounts WHERE is_active = 1')
        accounts = [r[0] for r in self.cursor.fetchall()]
        search_account_var = tk.StringVar()
        search_account_combo = ttk.Combobox(form_frame, textvariable=search_account_var, values=["সব"] + accounts, width=30, state='readonly')
        search_account_combo.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        search_account_combo.set("সব")

        tk.Label(form_frame, text="ধরন (ঐচ্ছিক):", font=('Segoe UI', 11, 'bold'), bg=self.colors['card']).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        search_type_var = tk.StringVar(value="সব")
        search_type_frame = tk.Frame(form_frame, bg=self.colors['card'])
        search_type_frame.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        tk.Radiobutton(search_type_frame, text="সব", variable=search_type_var, value="সব", bg=self.colors['card']).pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(search_type_frame, text="আয়", variable=search_type_var, value="আয়", bg=self.colors['card']).pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(search_type_frame, text="খরচ", variable=search_type_var, value="খরচ", bg=self.colors['card']).pack(side=tk.LEFT, padx=5)

        def perform_search():
            for item in search_results_tree.get_children():
                search_results_tree.delete(item)

            search_term = search_term_entry.get()
            account_filter = search_account_var.get()
            type_filter = search_type_var.get()

            query = '''
                SELECT t.transaction_date, t.transaction_time, a.account_name, a.account_type, t.amount, t.description, pm.method_name
                FROM transactions t
                JOIN accounts a ON t.account_id = a.account_id
                LEFT JOIN payment_methods pm ON t.payment_method_id = pm.method_id
                WHERE t.is_deleted = 0
            '''
            params = []

            if search_term:
                query += " AND (t.description LIKE ? OR a.account_name LIKE ?)"
                params.append(f'%{search_term}%')
                params.append(f'%{search_term}%')

            if account_filter != "সব":
                query += " AND a.account_name = ?"
                params.append(account_filter)

            if type_filter != "সব":
                query += " AND a.account_type = ?"
                params.append(type_filter)

            query += " ORDER BY t.transaction_date DESC, t.transaction_time DESC"

            self.cursor.execute(query, tuple(params))
            results = self.cursor.fetchall()

            for row in results:
                search_results_tree.insert('', 'end', values=(row[0], row[1], row[2], row[3], f"৳ {row[4]:,.2f}", row[5], row[6]))

        tk.Button(
            form_frame,
            text="অনুসন্ধান করুন",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=20,
            pady=10,
            bd=0,
            cursor='hand2',
            command=perform_search
        ).grid(row=3, column=0, columnspan=2, pady=10)

        # Search Results
        results_frame = tk.Frame(search_frame, bg=self.colors['card'])
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('তারিখ', 'সময়', 'খাত', 'ধরন', 'পরিমাণ', 'বিবরণ', 'পেমেন্ট পদ্ধতি')
        search_results_tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=10)
        for col in columns:
            search_results_tree.heading(col, text=col)
            search_results_tree.column(col, width=100)

        search_results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=search_results_tree.yview)
        search_results_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def show_settings(self):
        """Settings - সম্পূর্ণ"""
        self.clear_content_area()

        tk.Label(
            self.content_area,
            text="⚙️ সেটিংস",
            font=('Segoe UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(anchor='w', pady=(0, 20))

        settings_frame = tk.Frame(self.content_area, bg=self.colors['card'], relief=tk.RAISED, borderwidth=1)
        settings_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Password Management
        password_frame = tk.LabelFrame(settings_frame, text="পাসওয়ার্ড ব্যবস্থাপনা", bg=self.colors['card'], fg=self.colors['text'], font=('Segoe UI', 12, 'bold'))
        password_frame.pack(fill=tk.X, padx=10, pady=10, ipadx=5, ipady=5)

        tk.Label(password_frame, text="বর্তমান পাসওয়ার্ড:", bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        current_password_entry = tk.Entry(password_frame, show='●', width=30)
        current_password_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        tk.Label(password_frame, text="নতুন পাসওয়ার্ড:", bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        new_password_entry = tk.Entry(password_frame, show='●', width=30)
        new_password_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        tk.Label(password_frame, text="পাসওয়ার্ড নিশ্চিত করুন:", bg=self.colors['card']).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        confirm_password_entry = tk.Entry(password_frame, show='●', width=30)
        confirm_password_entry.grid(row=2, column=1, sticky='ew', padx=5, pady=5)

        def change_password():
            current_pass = current_password_entry.get()
            new_pass = new_password_entry.get()
            confirm_pass = confirm_password_entry.get()

            if not self.check_password_exists():
                # No password set yet, treat current_pass as empty
                if new_pass == confirm_pass and new_pass:
                    password_hash = hashlib.sha256(new_pass.encode()).hexdigest()
                    self.cursor.execute('INSERT OR REPLACE INTO security (id, password_hash) VALUES (1, ?)', (password_hash,))
                    self.conn.commit()
                    messagebox.showinfo("সফল", "পাসওয়ার্ড সফলভাবে সেট করা হয়েছে!")
                    current_password_entry.delete(0, tk.END)
                    new_password_entry.delete(0, tk.END)
                    confirm_password_entry.delete(0, tk.END)
                else:
                    messagebox.showerror("ত্রুটি", "নতুন পাসওয়ার্ড এবং নিশ্চিত পাসওয়ার্ড মিলছে না বা খালি আছে!")
                return

            # Existing password change
            current_pass_hash = hashlib.sha256(current_pass.encode()).hexdigest()
            self.cursor.execute('SELECT password_hash FROM security WHERE id = 1')
            stored_hash = self.cursor.fetchone()[0]

            if current_pass_hash != stored_hash:
                messagebox.showerror("ত্রুটি", "বর্তমান পাসওয়ার্ড ভুল!")
                return

            if new_pass != confirm_pass:
                messagebox.showerror("ত্রুটি", "নতুন পাসওয়ার্ড এবং নিশ্চিত পাসওয়ার্ড মিলছে না!")
                return

            if not new_pass:
                messagebox.showerror("ত্রুটি", "নতুন পাসওয়ার্ড খালি হতে পারে না!")
                return

            new_pass_hash = hashlib.sha256(new_pass.encode()).hexdigest()
            self.cursor.execute('UPDATE security SET password_hash = ? WHERE id = 1', (new_pass_hash,))
            self.conn.commit()
            messagebox.showinfo("সফল", "পাসওয়ার্ড সফলভাবে পরিবর্তন করা হয়েছে!")
            current_password_entry.delete(0, tk.END)
            new_password_entry.delete(0, tk.END)
            confirm_password_entry.delete(0, tk.END)

        tk.Button(
            password_frame,
            text="পাসওয়ার্ড পরিবর্তন করুন",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=change_password
        ).grid(row=3, column=0, columnspan=2, pady=10)

        # Security Question (for password recovery)
        security_q_frame = tk.LabelFrame(settings_frame, text="নিরাপত্তা প্রশ্ন সেট করুন", bg=self.colors['card'], fg=self.colors['text'], font=('Segoe UI', 12, 'bold'))
        security_q_frame.pack(fill=tk.X, padx=10, pady=10, ipadx=5, ipady=5)

        tk.Label(security_q_frame, text="নিরাপত্তা প্রশ্ন:", bg=self.colors['card']).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        security_question_entry = tk.Entry(security_q_frame, width=40)
        security_question_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        tk.Label(security_q_frame, text="আপনার উত্তর:", bg=self.colors['card']).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        security_answer_entry = tk.Entry(security_q_frame, show='●', width=40)
        security_answer_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        def set_security_question():
            question = security_question_entry.get()
            answer = security_answer_entry.get()

            if not question or not answer:
                messagebox.showerror("ত্রুটি", "প্রশ্ন এবং উত্তর উভয়ই পূরণ করুন!")
                return

            answer_hash = hashlib.sha256(answer.encode()).hexdigest()
            self.cursor.execute('UPDATE security SET security_question = ?, security_answer_hash = ? WHERE id = 1', (question, answer_hash))
            self.conn.commit()
            messagebox.showinfo("সফল", "নিরাপত্তা প্রশ্ন সফলভাবে সেট করা হয়েছে!")
            security_question_entry.delete(0, tk.END)
            security_answer_entry.delete(0, tk.END)

        tk.Button(
            security_q_frame,
            text="নিরাপত্তা প্রশ্ন সেট করুন",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=set_security_question
        ).grid(row=2, column=0, columnspan=2, pady=10)

        # Cloud Synchronization
        cloud_sync_frame = tk.LabelFrame(settings_frame, text="ক্লাউড সিঙ্ক্রোনাইজেশন", bg=self.colors['card'], fg=self.colors['text'], font=('Segoe UI', 12, 'bold'))
        cloud_sync_frame.pack(fill=tk.X, padx=10, pady=10, ipadx=5, ipady=5)

        tk.Button(
            cloud_sync_frame,
            text="Google Drive-এ আপলোড করুন",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=self._upload_to_drive
        ).pack(pady=5)

        tk.Button(
            cloud_sync_frame,
            text="Google Drive থেকে ডাউনলোড করুন",
            font=('Segoe UI', 11, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            padx=15,
            pady=8,
            bd=0,
            cursor='hand2',
            command=self._download_from_drive
        ).pack(pady=5)

        # Theme Selection
        theme_frame = tk.LabelFrame(settings_frame, text="থিম সেটিংস", bg=self.colors['card'], fg=self.colors['text'], font=('Segoe UI', 12, 'bold'))
        theme_frame.pack(fill=tk.X, padx=10, pady=10, ipadx=5, ipady=5)

        tk.Label(theme_frame, text="থিম নির্বাচন করুন:", bg=self.colors['card']).pack(side=tk.LEFT, padx=5, pady=5)
        theme_var = tk.StringVar(value=self.current_theme)
        tk.Radiobutton(theme_frame, text="Light", variable=theme_var, value="light", bg=self.colors['card'], command=lambda: self.apply_theme(theme_var.get())).pack(side=tk.LEFT, padx=5)
        # Add more themes here if needed

            security_question_entry.delete(0, tk.END)
            security_answer_entry.delete(0, tk.END)

    def _upload_to_drive(self):
        if self.drive_sync.service:
            try:
                self.conn.close() # Close DB connection before upload
                if self.drive_sync.upload_db_file():
                    messagebox.showinfo("সফল", "ডেটা Google Drive-এ সফলভাবে আপলোড করা হয়েছে!")
                else:
                    messagebox.showerror("ত্রুটি", "Google Drive-এ ডেটা আপলোড করতে ব্যর্থ।")
            except Exception as e:
                self.logger.error("Error uploading to Google Drive", exc_info=True)
                messagebox.showerror("ত্রুটি", f"Google Drive-এ আপলোড করতে ব্যর্থ: {e}")
            finally:
                self.init_database() # Re-open DB connection
        else:
            messagebox.showerror("ত্রুটি", "Google Drive পরিষেবা উপলব্ধ নয়। 'credentials.json' ফাইলটি সঠিক আছে কিনা নিশ্চিত করুন।")

    def _download_from_drive(self):
        if self.drive_sync.service:
            if messagebox.askyesno("নিশ্চিত করুন", "Google Drive থেকে ডেটা ডাউনলোড করলে আপনার বর্তমান স্থানীয় ডেটা ওভাররাইট হয়ে যাবে। আপনি কি নিশ্চিত?"):
                try:
                    self.conn.close() # Close DB connection before download
                    if self.drive_sync.download_db_file():
                        messagebox.showinfo("সফল", "ডেটা Google Drive থেকে সফলভাবে ডাউনলোড করা হয়েছে!")
                        self.init_database() # Re-initialize DB connection with new file
                        self.load_dashboard() # Refresh UI
                    else:
                        messagebox.showerror("ত্রুটি", "Google Drive থেকে ডেটা ডাউনলোড করতে ব্যর্থ। ফাইলটি Drive-এ নাও থাকতে পারে।")
                except Exception as e:
                    self.logger.error("Error downloading from Google Drive", exc_info=True)
                    messagebox.showerror("ত্রুটি", f"Google Drive থেকে ডাউনলোড করতে ব্যর্থ: {e}")
                finally:
                    self.init_database() # Ensure DB connection is open
        else:
            messagebox.showerror("ত্রুটি", "Google Drive পরিষেবা উপলব্ধ নয়। 'credentials.json' ফাইলটি সঠিক আছে কিনা নিশ্চিত করুন।")

    def apply_theme(self, theme_name):
        if theme_name in self.themes:
            self.current_theme = theme_name
            self.colors = self.themes[self.current_theme]
            
            # Update root and main container
            self.root.configure(bg=self.colors['bg'])
            self.main_container.configure(bg=self.colors['bg'])
            self.content_area.configure(bg=self.colors['bg'])

            # Update nav bar
            self.nav_bar.configure(bg=self.colors['primary'])
            for widget in self.nav_bar.winfo_children():
                if isinstance(widget, tk.Label):
                    widget.configure(bg=self.colors['primary'])

            # Update sidebar
            self.sidebar.configure(bg=self.colors['card'])
            self.menu_label.configure(bg=self.colors['card'], fg=self.colors['text'])
            for btn in self.menu_buttons:
                btn.configure(bg=self.colors['card'], fg=self.colors['text'])
                btn.bind('<Enter>', lambda e, b=btn: b.config(bg=self.colors['primary'], fg='white'))
                btn.bind('<Leave>', lambda e, b=btn: b.config(bg=self.colors['card'], fg=self.colors['text']))

            # Reload current view to apply theme to its widgets
            # This is a simplified approach; a full dynamic theme would require
            # updating all widgets in the current view.
            # For now, we'll just reload the dashboard as an example.
            self.load_dashboard()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    root = tk.Tk()
    app = AdvancedMoneyTracker(root)
    app.run()
